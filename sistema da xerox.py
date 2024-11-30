import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Protection
from datetime import datetime

# Função que será chamada quando o botão for clicado para calcular a renda
def calcular_renda():
    try:
        c = float(entry_copias.get()) if entry_copias.get() else 0.0
        ik = float(entry_imp_koycera.get()) if entry_imp_koycera.get() else 0.0
        ib = float(entry_imp_brother.get()) if entry_imp_brother.get() else 0.0
        s = float(entry_scaners.get()) if entry_scaners.get() else 0.0
        pix = float(entry_pix.get()) if entry_pix.get() else 0.0
        pbi = int(entry_pbi.get()) if entry_pbi.get() else 0
        pki = int(entry_pki.get()) if entry_pki.get() else 0
        pkc = int(entry_pkc.get()) if entry_pkc.get() else 0
        pks = int(entry_pks.get()) if entry_pks.get() else 0
        registro_pix = text_registpix.get("1.0", "end-1c").strip()
        obs = text_obs.get("1.0", "end-1c").strip()
        
        i = ik + ib
        p = (pbi + pki + pkc)
        ps = pks * 0.5
        perd = p * 0.2 + ps
        data = datetime.now()
        data_atual = data.strftime("%d/%m/%Y")

        with open("C:\Sistem\dados.txt", "r") as dados:
            lista = [x.strip() for x in dados]
        
        v1, v2, v3, v4 = float(lista[0]), float(lista[1]), float(lista[2]), float(lista[3])

        x = ik - v2
        y = ib - v3
        z = c - v1
        k = s - v4

        rik = x * 0.2
        rib = y * 0.2
        rc = z * 0.2
        rs = k * 0.5
        ri = rik + rib

        vpf = (ri + rc + rs) - perd
        ve = vpf - pix
        rend_tot = pix + ve

        # Atualiza os dados no arquivo
        with open("C://Sistem/dados.txt", "w") as dados:
            dados.write(f"{c}\n{ik}\n{ib}\n{s}")

        with open("C://Sistem/dados.txt", "r") as dados:
            lista = [x.strip() for x in dados]
        
        l1, l2, l3, l4 = float(lista[0]), float(lista[1]), float(lista[2]), float(lista[3])

        wb = load_workbook("C:/Sistem/planilha.xlsx")
        ws = wb.active

        ws['F15'] = l1
        ws['E15'] = v1
        ws['F16'] = l2
        ws['E16'] = v2
        ws['F17'] = l3
        ws['E17'] = v3
        ws['F18'] = l4
        ws['E18'] = v4
        ws['J22'] = pix
        ws['I11'] = data_atual
        ws['H15'] = pkc
        ws['H16'] = pki
        ws['H17'] = pbi
        ws['H18'] = pks
        ws['G23'] = registro_pix
        ws['D23'] = obs

        proteger_planilha_com_senha(wb, '//10.7.1.7/reprografia-cbpav/relatorio.xlsx', 'minhasenha')

        messagebox.showinfo("Resultado", f"Renda total: {rend_tot:.2f} reais\nRenda em espécie: {ve:.2f} reais\nBordearux disponível em:\nreprografia-cbpav")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

def proteger_planilha_com_senha(wb, output_path, senha):
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=False)
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.set_password(senha)
    wb.save(output_path)

# Função para mostrar a janela de alerta
def mostrar_alerta():
    janela_alerta = tk.Toplevel(window)
    janela_alerta.title("Atenção!")
    janela_alerta.geometry("500x200")
    janela_alerta.configure(bg="lightyellow")
    janela_alerta.transient(window)
    janela_alerta.grab_set()
    janela_alerta.focus_force()
    msg_alerta = "Atenção: Por favor, insira os valores corretamente. Valores incorretos podem comprometer o funcionamento do programa inteiro!"
    label_alerta = tk.Label(janela_alerta, text=msg_alerta, bg="lightyellow", font=("Arial", 12, "bold"), wraplength=400)
    label_alerta.pack(pady=20)
    btn_fechar = tk.Button(janela_alerta, text="Fechar", command=janela_alerta.destroy, bg="lightgreen")
    btn_fechar.pack(pady=10)
    janela_alerta.wait_window()

# Função para abrir a janela de redefinir valores
def janela_definir_novos_valores():
    janela_novos_valores = tk.Toplevel(window)
    janela_novos_valores.title("Definir Novos Valores iniciais")
    janela_novos_valores.geometry("400x400")
    janela_novos_valores.configure(bg="lightyellow")

    # Labels e entradas para redefinir os valores
    tk.Label(janela_novos_valores, text="Novo valor para Cópias Kyocera:", font=("Arial", 12)).pack(pady=10)
    entry_novo_valor_copias = tk.Entry(janela_novos_valores, font=("Arial", 12), relief="groove", borderwidth=2)
    entry_novo_valor_copias.pack(pady=5)

    tk.Label(janela_novos_valores, text="Novo valor para Impressões Kyocera:", font=("Arial", 12)).pack(pady=10)
    entry_novo_valor_imp_kyocera = tk.Entry(janela_novos_valores, font=("Arial", 12), relief="groove", borderwidth=2)
    entry_novo_valor_imp_kyocera.pack(pady=5)

    tk.Label(janela_novos_valores, text="Novo valor para Impressões Brother:", font=("Arial", 12)).pack(pady=10)
    entry_novo_valor_imp_brother = tk.Entry(janela_novos_valores, font=("Arial", 12), relief="groove", borderwidth=2)
    entry_novo_valor_imp_brother.pack(pady=5)

    tk.Label(janela_novos_valores, text="Novo valor para Scanners Kyocera:", font=("Arial", 12)).pack(pady=10)
    entry_novo_valor_scaners = tk.Entry(janela_novos_valores, font=("Arial", 12), relief="groove", borderwidth=2)
    entry_novo_valor_scaners.pack(pady=5)

    # Função para salvar os novos valores no arquivo
    def salvar_novos_valores():
        try:
            novo_valor_copias = entry_novo_valor_copias.get()
            novo_valor_imp_kyocera = entry_novo_valor_imp_kyocera.get()
            novo_valor_imp_brother = entry_novo_valor_imp_brother.get()
            novo_valor_scaners = entry_novo_valor_scaners.get()

            if not novo_valor_copias or not novo_valor_imp_kyocera or not novo_valor_imp_brother or not novo_valor_scaners:
                messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
                return

            # Grava os novos valores no arquivo
            with open("C://Sistem/dados.txt", "w") as dados:
                dados.write(f"{novo_valor_copias}\n{novo_valor_imp_kyocera}\n{novo_valor_imp_brother}\n{novo_valor_scaners}")

            messagebox.showinfo("Sucesso", "Novos valores definidos com sucesso!")
            janela_novos_valores.destroy()  # Fecha a janela de redefinir valores
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao salvar os novos valores: {e}")

    # Botão para salvar os novos valores
    btn_salvar_novos_valores = tk.Button(janela_novos_valores, text="Salvar Novos Valores", command=salvar_novos_valores, bg="lightgreen", font=("Arial", 12, "bold"))
    btn_salvar_novos_valores.pack(pady=20)

# Janela principal
window = tk.Tk()
window.title("Sistema da Xerox")
window.geometry("700x700")
window.configure(bg="#f0f0f0")
font_padrao = ("Arial", 12)

# Exibe a mensagem de alerta ao iniciar
mostrar_alerta()

# Layout e entradas da interface principal
campos = [
    ("Cópias Kyocera:", "lightblue"),
    ("Impressões Kyocera:", "white"),
    ("Scaners Kyocera:", "lightblue"),
    ("Impressões Brother:", "white"),
    ("Páginas perdidas (cópias Kyocera):", "lightblue"),
    ("Páginas perdidas (impressões Kyocera):", "white"),
    ("Páginas perdidas (scaners Kyocera):", "lightblue"),
    ("Páginas perdidas (impressões Brother):", "white"),
    ("Receita em Pix:", "lightblue")
]

entries = []
for i, (label_text, bg_color) in enumerate(campos):
    tk.Label(window, text=label_text, bg=bg_color, font=font_padrao, padx=10, pady=5).grid(row=i, column=0, sticky="e", padx=10, pady=5)
    entry = tk.Entry(window, font=font_padrao, relief="groove", borderwidth=2)
    entry.grid(row=i, column=1, pady=5, padx=10, sticky="w")
    entries.append(entry)

(entry_copias, entry_imp_koycera, entry_scaners, entry_imp_brother, entry_pkc, entry_pki, entry_pks, entry_pbi, entry_pix) = entries

tk.Label(window, text="Registros de Pix:", font=font_padrao).grid(row=9, column=0, pady=5)
text_registpix = tk.Text(window, height=4, width=40, relief="groove", borderwidth=2)
text_registpix.grid(row=9, column=1, pady=5, padx=10)

tk.Label(window, text="Observações:", font=font_padrao).grid(row=10, column=0, pady=5)
text_obs = tk.Text(window, height=4, width=40, relief="groove", borderwidth=2)
text_obs.grid(row=10, column=1, pady=5, padx=10)

btn_calcular = tk.Button(window, text="Calcular e Gerar Relatório", command=calcular_renda, bg="lightgreen", font=("Arial", 12, "bold"), relief="raised")
btn_calcular.grid(row=11, column=0, columnspan=2, pady=15)

# Botão para definir novos valores
btn_definir_novos_valores = tk.Button(window, text="Definir Novos Valores iniciais", command=janela_definir_novos_valores, bg="lightcoral", font=("Arial", 12, "bold"))
btn_definir_novos_valores.grid(row=12, column=0, columnspan=2, pady=15)

window.mainloop()
